"""Generate the offline catch-list web page and the Excel workbook."""
import json, os, sys
from pvpoke_data import build_all, CATS
import availability

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/outputs"
TIER_LABEL = {"core": "1 - Core", "flex": "2 - Flex", "niche": "3 - Deep"}


def write_html(data, path):
    tpl = open(os.path.join(HERE, "template.html"), encoding="utf-8").read()
    blob = json.dumps(data, separators=(",", ":")).replace("</", "<\\/")
    open(path, "w", encoding="utf-8").write(tpl.replace("__DATA__", blob))


def write_xlsx(data, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F = "Arial"
    head = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="BFBFBF")
    tier_fill = {"1 - Core": PatternFill("solid", fgColor="C6E0B4"),
                 "2 - Flex": PatternFill("solid", fgColor="FFE699"),
                 "3 - Deep": PatternFill("solid", fgColor="F2F2F2")}
    hdr = ["Tier", "Available Now", "Lure", "Source", "Catch This", "Becomes", "Shadow",
           "Overall", "Lead", "Switch", "Closer", "Best Role", "Roles in Top 75", "XL Needed"]
    widths = [11, 22, 28, 15, 22, 38, 13, 9, 8, 9, 9, 14, 9, 11]
    ctr = {1, 4, 7, 8, 9, 10, 11, 12, 13, 14}

    wb = Workbook()
    wb.remove(wb.active)
    for key, lg in data["leagues"].items():
        ws = wb.create_sheet(lg["label"].replace(" League", ""))
        ws.append(hdr)
        for r in lg["rows"]:
            now = r.get("now")
            lure = ", ".join(r.get("lures") or [])
            if r.get("lureEvo"):
                lure = (lure + " | " if lure else "") + "; ".join(r["lureEvo"])
            ws.append([TIER_LABEL[r["tier"]], now["label"] if now else "", lure, r["source"],
                       r["target"], ", ".join(r["becomes"]), r["avail"],
                       r["overall"] or "", r["lead"], r["switch"], r["closer"],
                       r["bestRole"], r["roles75"], "Yes" if r["xl"] else "No"])
        for c in ws[1]:
            c.font = Font(name=F, bold=True, color="FFFFFF", size=11)
            c.fill = head
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(hdr)):
            for j, c in enumerate(row, 1):
                c.font = Font(name=F, size=11)
                c.border = Border(bottom=thin)
                c.alignment = Alignment(horizontal="center" if j in ctr else "left")
            row[0].fill = tier_fill[row[0].value]
            if row[1].value:
                row[1].font = Font(name=F, size=11, bold=True, color="1F3864")
            if row[6].value == "Shadow only":
                row[6].font = Font(name=F, size=11, bold=True, color="7030A0")
            if row[2].value and "evolve" in str(row[2].value):
                row[2].font = Font(name=F, size=11, bold=True, color="0B5A78")
            for j in (8, 9, 10, 11):
                v = row[j - 1].value
                if isinstance(v, int) and v <= 75:
                    row[j - 1].font = Font(name=F, size=11, bold=True)
            if row[13].value == "Yes":
                row[13].font = Font(name=F, size=11, color="C00000")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{ws.max_row}"
        ws.row_dimensions[1].height = 30

    ws = wb.create_sheet("Notes")
    for r in [["Field", "Detail"],
              ["Tier 1 - Core", "Top 25 in at least one of Overall / Lead / Switch / Closer."],
              ["Tier 2 - Flex", "Top 75 in two or more categories, no top-25 finish."],
              ["Tier 3 - Deep", "Top 75 in exactly one category."],
              ["Source", "Wild = catchable while walking. Raid, Eggs only, Special Research, Mystery Box = not."],
              ["Available Now", "Event spawn, raid rotation, or egg pool that is live today, with its end date where one is published. Blank means not currently featured -- it may still spawn normally in your area."],
              ["Lure", "Which Lure Modules boost this Pokemon's types out of your local spawn pool, plus any Lure required to evolve it. Glacial = Water/Ice. Mossy = Bug/Grass/Poison. Magnetic = Electric/Steel/Rock. Rainy = Water/Bug/Electric."],
              ["Shadow", "Normal only = the wild catch is the meta piece. Both = wild catch works and the Shadow is also ranked. Shadow only = candy farming only; the usable Pokemon comes from a Rocket battle."],
              ["Rank columns", "Best rank the family reaches in that category, counting Shadow and non-Shadow forms together, since they share a candy pool."],
              ["XL Needed", "Great and Ultra: cannot reach the CP cap at level 40 with perfect IVs. Master is uncapped, so XL is always a real power gap."],
              ["Source data", f"PvPoke All Pokemon rankings, gamemaster dated {data['gamemaster']}. Generated {data['generated']}."]]:
        ws.append(r)
    for c in ws[1]:
        c.font = Font(name=F, bold=True, color="FFFFFF")
        c.fill = head
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for c in row:
            c.font = Font(name=F, size=11)
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100
    wb.save(path)


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    data = build_all()
    print("checking current availability...")
    avail = availability.build()
    data["availability"] = len(avail)
    for lg in data["leagues"].values():
        availability.attach(lg["rows"], avail)
    write_html(data, os.path.join(OUT, "catch-list.html"))
    write_xlsx(data, os.path.join(OUT, "catch-list.xlsx"))
    for k, v in data["leagues"].items():
        n = lambda t: sum(1 for r in v["rows"] if r["tier"] == t)
        print(f"{v['label']}: {len(v['rows'])} targets ({n('core')} core, {n('flex')} flex, {n('niche')} deep)")
    print("gamemaster", data["gamemaster"])
